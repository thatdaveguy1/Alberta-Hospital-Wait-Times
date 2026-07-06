#!/usr/bin/env python3
"""
Stream-extract attachment-rate rows from the CIHI Indicator Library workbook.

The workbook is a single 822K-row sheet ("Sheet1") with 33 columns. SheetJS
cannot load it (813 MB uncompressed), so we use openpyxl in read_only mode.

Outputs JSON to stdout: { "rates": [...], "error": null }
Each rate: { id, geography, demographic_group, reporting_year, metric_value }
"""
import json
import re
import sys

import openpyxl

WORKBOOK_PATH = sys.argv[1] if len(sys.argv) > 1 else "/tmp/cihi-indicator-library.xlsx"

KEEP_GEOGRAPHIES = {
    "alberta": "Alberta",
    "canada": "Canada",
    "calgary zone": "Calgary Zone",
    "edmonton zone": "Edmonton Zone",
    "central zone": "Central Zone",
    "north zone": "North Zone",
    "south zone": "South Zone",
    "calgary zone alta": "Calgary Zone",
    "edmonton zone alta": "Edmonton Zone",
    "central zone alta": "Central Zone",
    "north zone alta": "North Zone",
    "south zone alta": "South Zone",
    "calgary": "Calgary Zone",
    "edmonton": "Edmonton Zone",
    "central": "Central Zone",
    "north": "North Zone",
    "south": "South Zone",
}

INDICATOR_FRAGMENTS = [
    "regular health provider",
    "regular health care provider",
    "regular provider",
    "attachment",
]


def normalize(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]", " ", str(s).lower())).strip()


def find_col(headers, fragments):
    for i, h in enumerate(headers):
        norm = normalize(h)
        if all(normalize(f) in norm for f in fragments):
            return i
    return -1


def derive_demographic(breakdown, breakdown_val, segment_value):
    """Map CIHI breakdown columns to a demographic-group label.

    When Level 1 breakdown is "Not applicable", the row is an overall rate
    for the segment (Adults or Children and youth). We map the Adults overall
    rate to "All Residents" since the dashboard uses that as the headline.
    """
    bd = normalize(breakdown)
    bv = normalize(breakdown_val)
    seg = normalize(segment_value)

    if bd == "not applicable" or not bv or bv == "not applicable":
        if "children" in seg:
            return "Children & Youth (0-17)"
        # Adults overall rate → headline "All Residents"
        return "All Residents"

    if "age group" in bd:
        if "18 to 34" in bv or "35 to 49" in bv or "50 to 64" in bv:
            return "Adults (18-64)"
        if "65" in bv:
            return "Seniors (65+)"
        if "12 to 17" in bv or "5 to 11" in bv or "1 to 4" in bv:
            return "Children & Youth (0-17)"
        return str(breakdown_val) if breakdown_val else "All Residents"

    if "income" in bd:
        if "q1" in bv or "lowest" in bv:
            return "Lowest Income Quintile"
        if "q5" in bv or "highest" in bv:
            return "Highest Income Quintile"
        if "q2" in bv:
            return "Lower-Middle Income Quintile"
        if "q3" in bv:
            return "Middle Income Quintile"
        if "q4" in bv:
            return "Upper-Middle Income Quintile"
        return str(breakdown_val) if breakdown_val else "All Residents"

    if "urban" in bd and "rural" in bd:
        if "rural" in bv:
            return "Rural / Remote Areas"
        if "urban" in bv:
            return "Urban Centres"
        return str(breakdown_val) if breakdown_val else "All Residents"

    if "sex" in bd or "gender" in bd:
        # Map specific gender/sex values
        if "female" in bv:
            return "Female"
        if "male" in bv:
            return "Male"
        if "girls" in bv:
            return "Girls"
        if "boys" in bv:
            return "Boys"
        return str(breakdown_val) if breakdown_val else "All Residents"

    return str(breakdown_val) if breakdown_val else "All Residents"


def main():
    try:
        wb = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
        sheet = wb["Sheet1"]

        header_iter = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        header_row = next(header_iter, None)
        if header_row is None:
            print(json.dumps({"rates": [], "error": "No header row found"}))
            return

        headers = [str(h) if h is not None else "" for h in header_row]

        indicator_col = find_col(headers, ["indicator"])
        geo_col = find_col(headers, ["place or organization"])
        value_col = find_col(headers, ["metric value"])
        year_col = find_col(headers, ["time frame"])
        main_metric_col = find_col(headers, ["main metric"])
        breakdown_col = find_col(headers, ["level 1 breakdown"])
        breakdown_val_col = find_col(headers, ["level 1 breakdown value"])
        segment_val_col = find_col(headers, ["segment value"])

        if indicator_col < 0 or geo_col < 0 or value_col < 0:
            print(json.dumps({
                "rates": [],
                "error": f"Required columns not found (indicator={indicator_col}, geo={geo_col}, value={value_col})",
            }))
            return

        rates = []
        seen_ids = set()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            indicator = row[indicator_col] if indicator_col < len(row) else None
            if indicator is None:
                continue
            norm_ind = normalize(indicator)
            if not any(normalize(f) in norm_ind for f in INDICATOR_FRAGMENTS):
                continue

            if main_metric_col >= 0 and main_metric_col < len(row):
                main_metric = row[main_metric_col]
                if main_metric is not None and normalize(main_metric) != "yes":
                    continue

            geo_raw = row[geo_col] if geo_col < len(row) else None
            if geo_raw is None:
                continue
            geo_norm = normalize(geo_raw)
            geography = KEEP_GEOGRAPHIES.get(geo_norm)
            if geography is None:
                continue

            value = row[value_col] if value_col < len(row) else None
            if value is None or value == "":
                continue
            try:
                metric_value = float(value)
            except (ValueError, TypeError):
                continue

            year = row[year_col] if year_col >= 0 and year_col < len(row) else None
            year_str = str(year) if year is not None else ""
            m = re.match(r"^(\d{4})", year_str)
            if not m:
                continue
            reporting_year = m.group(1)

            breakdown = row[breakdown_col] if breakdown_col >= 0 and breakdown_col < len(row) else None
            breakdown_val = row[breakdown_val_col] if breakdown_val_col >= 0 and breakdown_val_col < len(row) else None
            segment_val = row[segment_val_col] if segment_val_col >= 0 and segment_val_col < len(row) else None
            demographic = derive_demographic(breakdown, breakdown_val, segment_val)

            rate_id = f"att_{normalize(geography).replace(' ', '_')}_{normalize(demographic).replace(' ', '_')}_{reporting_year}"
            if rate_id in seen_ids:
                continue
            seen_ids.add(rate_id)

            rates.append({
                "id": rate_id,
                "geography": geography,
                "demographic_group": demographic,
                "reporting_year": reporting_year,
                "metric_value": metric_value,
            })

        wb.close()
        print(json.dumps({"rates": rates, "error": None}))

    except Exception as e:
        print(json.dumps({"rates": [], "error": str(e)}))


if __name__ == "__main__":
    main()
